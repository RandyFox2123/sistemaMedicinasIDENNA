
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect, HttpResponse, Http404
from django.urls import reverse
from django.core.paginator import Paginator
from django.views.decorators.http import require_safe, require_http_methods, require_POST
from openpyxl import Workbook
from  . models import Medicina, Ubicacion, Presentacion_Medicamento
import os
import logging
from django.conf import settings
from decimal import Decimal
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from datetime import date
from django.utils.http import urlencode
from django.conf import settings
from django.db.models import Case, When, Value, BooleanField
from django.utils import timezone
from datetime import datetime
from datetime import timedelta
from django.db.models import Q
from .forms import MedicinaForm



logger = logging.getLogger(__name__)

error_no_encontrado = 'Lo sentimos, el recurso que solicitas no fue encontrado'
error_inesperado = 'Lo sentimos, pero ha ocurrido un error inesperado en el proceso'
metodo_no_permitido = 'Lo sentimos, hubo un error, el método no está permitido, no se puede proceder por seguridad'

class error_contexto(Exception):
    pass

def manejo_errores(vista_recibida):
    def wrapper(request, *args, **kwargs):
        mensaje_error = None
        try:
            return vista_recibida(request, *args, **kwargs)

        except TypeError as e:
            mensaje_error = 'Lo sentimos, hubo un error interno en el proceso'
            logger.error(f'!!! ERROR OBTENIDO !!!:\n\n{e}')
            return render(request, 'error.html', {'error_obtenido': mensaje_error})

        except Http404 as e:
            mensaje_error = error_no_encontrado
            logger.error(f'!!! ERROR OBTENIDO !!!:\n\n{e}')
            return render(request, 'error.html', {'error_obtenido': mensaje_error})

        except Exception as error_obtenido:
            if isinstance(error_obtenido, error_contexto):
                mensaje_error = str(error_obtenido)
                logger.error(mensaje_error)
            else:
                mensaje_error = error_inesperado
                logger.error(f'{error_obtenido}')
            return render(request, 'error.html', {'error_obtenido': mensaje_error})

    return wrapper



@manejo_errores
@require_http_methods(["GET", "POST"])
def login_ingreso(request):
    contexto = {"mensaje" : None, "perfil" : None}
    
    if request.method == "GET":
        return render(request, 'registration/login.html')
    
    else:
        perfil_ingresado = request.POST.get("username", None)
        contrasena_ingresada = request.POST.get("password", None)
        
        if not perfil_ingresado or not contrasena_ingresada:
            contexto["mensaje"] = "Debes ingresar usuario y contraseña"
            contexto["perfil"] = perfil_ingresado
            return render(request, 'registration/login.html', contexto)
        
        try:
            perfil = User.objects.get(username=perfil_ingresado)
            if perfil:
                if perfil.is_active == False:
                    contexto["mensaje"] = "El perfil existe, pero está desactivado pongase en contacto con un usuario superior para que active este perfil."
                    contexto["perfil"] = perfil_ingresado
                    return render(request, 'registration/login.html', contexto)
            
                else:
                    autenticacion = authenticate(username=perfil_ingresado, password=contrasena_ingresada)
                    if autenticacion:
                        print("Ingreso exitoso")
                        login(request, perfil)
                        return redirect("panel_principal")
                    
                    else:
                        contexto["mensaje"] = "El perfil existe pero la contraseña introducida es invalida." 
                        contexto["perfil"] = perfil_ingresado
                        return render(request, 'registration/login.html', contexto)
    
        except  User.DoesNotExist:
            contexto["mensaje"] = "Lo sentimos, este perfil no existe."
            contexto["perfil"] = perfil_ingresado
            return render(request, 'registration/login.html', contexto)
            
             
@manejo_errores
@require_safe
def cerrar_seccion(request):
    logout(request)
    return redirect("/")


@manejo_errores
def index(request):
    return render(request, 'index.html')



@manejo_errores
@require_safe
@login_required
def panel_principal(request):
    ubicaciones_activas = Ubicacion.objects.filter(estado=True)
    presentaciones_activas = Presentacion_Medicamento.objects.all()

    filtro_medicina = request.GET.get('filtro_medicina', '').strip()
    filtro_ubicacion_id = request.GET.get('filtro_ubicacion', '').strip()
    filtro_presentacion_id = request.GET.get('filtro_presentacion', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    page_num = request.GET.get('page', 1)

    medicinas_qs = Medicina.objects.filter(ubicacion__estado=True).annotate(
        estado_caducado_sql=Case(
            When(
                fecha_caducidad__isnull=False,
                fecha_caducidad__lte=timezone.now().date(),
                then=Value(True)
            ),
            default=Value(False),
            output_field=BooleanField()
        )
    )

    if filtro_medicina:
        medicinas_qs = medicinas_qs.filter(medicina__icontains=filtro_medicina)
    if filtro_ubicacion_id.isdigit():
        medicinas_qs = medicinas_qs.filter(ubicacion_id=int(filtro_ubicacion_id))
    if filtro_presentacion_id.isdigit():
        medicinas_qs = medicinas_qs.filter(presentacion_id=int(filtro_presentacion_id))
    if fecha_desde:
        medicinas_qs = medicinas_qs.filter(fecha_registro__gte=fecha_desde)
    if fecha_hasta:
        medicinas_qs = medicinas_qs.filter(fecha_registro__lte=fecha_hasta)

    medicinas_qs = medicinas_qs.order_by('-id_medicina')
    
    paginator = Paginator(medicinas_qs, 3)
    page_obj = paginator.get_page(page_num)

    for medicina in page_obj:
        medicina.caducir()

    for medicina in page_obj:
        if medicina.imagen_medicina and medicina.imagen_medicina.name:
            ruta_imagen = os.path.join(settings.MEDIA_ROOT, medicina.imagen_medicina.name)
            medicina.imagen_existe = os.path.isfile(ruta_imagen)
        else:
            medicina.imagen_existe = False

    get_params = request.GET.copy()
    if 'page' in get_params:
        get_params.pop('page')
    filtros_qs = get_params.urlencode()

    contexto = {
        'page_obj': page_obj,
        'ubicaciones_activas': ubicaciones_activas,
        'presentaciones_activas': presentaciones_activas,
        'filtro_medicina': filtro_medicina,
        'filtro_ubicacion_id': filtro_ubicacion_id,
        'filtro_presentacion_id': filtro_presentacion_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'filtros_qs': filtros_qs,
    }
    return render(request, 'panel_principal.html', contexto)



@manejo_errores
@login_required
@require_safe
def ver_medicina(request, id_medicina):
    medicina_consultada = get_object_or_404(Medicina, pk=id_medicina)
    medicina_consultada.caducir()
    ubicaciones_activas = Ubicacion.objects.filter(estado=True)
    presentaciones = Presentacion_Medicamento.objects.all()

    contexto = {
        "medicina_consultada": medicina_consultada, 
        "ubicaciones": ubicaciones_activas, 
        "presentaciones": presentaciones
    }
    return render(request, 'ver_medicina.html', contexto)


@manejo_errores
@login_required
@require_http_methods(["GET", "POST"])
def registrar_medicina(request):
    ubicaciones_activas = Ubicacion.objects.filter(estado=True)
    presentaciones = Presentacion_Medicamento.objects.all()
    
    if request.method == 'POST':
        formulario = MedicinaForm(request.POST, request.FILES)
        if formulario.is_valid():
            medicina = formulario.save(commit=False)
            medicina.creador_del_registro = request.user.username
            medicina.historial_edicion = 'Nadie'
            medicina.fecha_registro = date.today()
            medicina.save()
            medicina.caducir()
            return redirect('panel_principal')
    else:
        formulario = MedicinaForm()
    
    return render(request, 'registrar_medicina.html', {
        'form': formulario,
        'ubicaciones_activas': ubicaciones_activas, 
        'presentaciones': presentaciones
    })


@manejo_errores
@login_required
@require_http_methods(["GET", "POST"])
def editar_medicina(request, id_medicina):
    medicina_obj = get_object_or_404(Medicina, pk=id_medicina)
    medicina_obj.caducir()
    ubicaciones = Ubicacion.objects.filter(estado=True)
    presentaciones = Presentacion_Medicamento.objects.all()
    
    if request.method == 'POST':
        form = MedicinaForm(request.POST, request.FILES, instance=medicina_obj) 
        if form.is_valid():
            medicina = form.save(commit=False)
            nuevo_historial_edicion = f"{date.today()} : {request.user.username}"
            medicina.historial_edicion = medicina.historial_edicion + f", {nuevo_historial_edicion}"
            
            medicina.save()
            medicina.caducir()
            
            url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
            return HttpResponseRedirect(url_redireccion)
    else:
        form = MedicinaForm(instance=medicina_obj)
        
    contexto = {
        'form': form, 
        'medicina_obj': medicina_obj,
        'ubicaciones_activas': ubicaciones,
        'presentaciones': presentaciones,
    }
    return render(request, 'editar_medicina.html', contexto)



@manejo_errores
@login_required
@require_POST
def borrar_medicina(request, id_medicina):
    medicina_obj = get_object_or_404(Medicina, pk=id_medicina)
    medicina_obj.delete()

    url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
    return HttpResponseRedirect(url_redireccion)


@manejo_errores
@login_required
@require_POST
def sumar_cantidad_medicina(request, id_medicina):
    medicina = get_object_or_404(Medicina, pk=id_medicina)
    try:
        cantidad_sumar = request.POST.get('suma', '0').replace(',', '.')
        cantidad_sumar = Decimal(cantidad_sumar)

        if cantidad_sumar < 0:
            raise error_contexto("La cantidad a sumar no puede ser negativa.")
    except (ValueError, TypeError):
        raise error_contexto("Cantidad invalida.")

    medicina.cantidad += cantidad_sumar
    medicina.save()

    url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
    return HttpResponseRedirect(url_redireccion)


@manejo_errores
@login_required
@require_POST
def restar_cantidad_medicina(request, id_medicina):
    medicina = get_object_or_404(Medicina, pk=id_medicina)
    try:
        cantidad_restar = request.POST.get('resta', '0').replace(',', '.')
        cantidad_restar = Decimal(cantidad_restar)

        if cantidad_restar < 0:
            raise error_contexto("La cantidad a restar no puede ser negativa.")

    except (ValueError, TypeError):
        raise error_contexto("Cantidad invalida.")

    nueva_cantidad = max(medicina.cantidad - cantidad_restar, Decimal('0'))
    medicina.cantidad = nueva_cantidad
    medicina.save()

    url_redireccion = request.session.get('url_panel_principal', reverse('panel_principal'))
    return HttpResponseRedirect(url_redireccion)


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger



@manejo_errores
@login_required
@require_safe
def caducidades(request):
    hoy = timezone.now().date()
    en_30_dias = hoy + timedelta(days=30)
    hoy_str = str(hoy)
    medicinas_caducadas = Medicina.objects.filter(caducado=True).count()

    query = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)

    medicinas_proximas_a_caducar = Medicina.objects.filter(
        Q(caducado__isnull=True) | Q(caducado=False),
        fecha_caducidad__isnull=False,             
        fecha_caducidad__range=[hoy, en_30_dias]
    ).order_by('fecha_caducidad')

    if query:
        medicinas_proximas_a_caducar = medicinas_proximas_a_caducar.filter(
            medicina__icontains=query
        )


    paginator = Paginator(medicinas_proximas_a_caducar, 6)
    try:
        medicinas_page = paginator.page(page_number)
    except PageNotAnInteger:
        medicinas_page = paginator.page(1)
    except EmptyPage:
        medicinas_page = paginator.page(paginator.num_pages)

    for med in medicinas_page:
        if med.fecha_caducidad:
            med.dias_restantes = (med.fecha_caducidad - hoy).days
        else:
            med.dias_restantes = None

    cantidad_medicinas_proximas_a_caducar = paginator.count

    contexto = {
        'medicinas_caducadas': medicinas_caducadas,
        'cantidad_medicinas_proximas_a_caducar': cantidad_medicinas_proximas_a_caducar,
        'medicinas_proximas_a_caducar': medicinas_page,
        'query': query,
        'paginator': paginator,
        'page_obj': medicinas_page,
        'current_page': int(page_number),
        'total_pages': paginator.num_pages,
    }
    return render(request, "caducidades.html", contexto)


@manejo_errores
@login_required
@require_safe
def generar_excel(request):
    filtro_medicina = request.GET.get('filtro_medicina', '').strip()
    filtro_ubicacion_id = request.GET.get('filtro_ubicacion', '').strip()
    filtro_presentacion_id = request.GET.get('filtro_presentacion', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    medicinas_qs = Medicina.objects.filter(ubicacion__estado=True)

    if filtro_medicina:
        medicinas_qs = medicinas_qs.filter(medicina__icontains=filtro_medicina)

    if filtro_ubicacion_id.isdigit():
        medicinas_qs = medicinas_qs.filter(ubicacion_id=int(filtro_ubicacion_id))

    if filtro_presentacion_id.isdigit():
        medicinas_qs = medicinas_qs.filter(presentacion_id=int(filtro_presentacion_id))

    if fecha_desde:
        medicinas_qs = medicinas_qs.filter(fecha_registro__gte=fecha_desde)
    if fecha_hasta:
        medicinas_qs = medicinas_qs.filter(fecha_registro__lte=fecha_hasta)

    medicinas_qs = medicinas_qs.order_by('id_medicina')

    wb = Workbook()
    ws = wb.active
    ws.title = "Medicinas"

    ws['A1'] = 'Nombre medicina'
    ws['B1'] = 'Presentación'
    ws['C1'] = 'Cantidad'
    ws['D1'] = 'Laboratorio'
    ws['E1'] = 'Ubicación'
    ws['F1'] = 'Anaquel'
    ws['G1'] = 'Descripción'
    ws['H1'] = 'Observaciones'
    ws['I1'] = 'Creador del registro'
    ws['J1'] = 'Historial de edición'
    ws['K1'] = 'Fecha de registro'
    ws['M1'] = 'Fecha de caducidad'
    ws['N1'] = 'Caducado'

    row = 2
    for medicina in medicinas_qs:
        ws[f'A{row}'] = medicina.medicina
        ws[f'B{row}'] = medicina.presentacion.desc_presentacion_medicamento if medicina.presentacion else "Sin presentación"
        ws[f'C{row}'] = medicina.cantidad
        ws[f'D{row}'] = medicina.laboratorio or "No indica"
        ws[f'E{row}'] = medicina.ubicacion.desc_ubicacion if medicina.ubicacion else "Sin ubicación"
        ws[f'F{row}'] = medicina.anaquel
        ws[f'G{row}'] = medicina.descripcion
        ws[f'H{row}'] = medicina.observaciones or ""
        ws[f'I{row}'] = medicina.creador_del_registro
        ws[f'J{row}'] = medicina.historial_edicion
        ws[f'K{row}'] = medicina.fecha_registro.strftime('%Y-%m-%d') if medicina.fecha_registro else ""
        ws[f'M{row}'] = medicina.fecha_caducidad.strftime('%Y-%m-%d') if medicina.fecha_registro else ""

        if medicina.caducado == False:
            ws[f'N{row}'] = "No"
        elif medicina.caducado == True:
            ws[f'N{row}'] = "Si"
        else:
            ws[f'N{row}'] = "NO INDICADO"

        row += 1

    nombre_archivo = "Reporte_Medicinas.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response